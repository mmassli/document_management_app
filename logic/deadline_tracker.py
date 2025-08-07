"""
Deadline tracking functionality for the application.
Handles reminders, Excel generation, and email sending for department deadlines.
"""

import os
import json
import yaml
from datetime import datetime, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import messagebox, ttk
import pandas as pd
import tempfile
import platform
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import re


class DeadlineTracker:
    """Handles deadline tracking, reminders, and email sending"""
    
    def __init__(self, app):
        self.app = app
        self.tracking_file = Path("config/fristen_tracking.json")
        self.config_file = Path("config/fristen_config.yaml")
        self.config = self._load_config()
        
    def _load_config(self):
        """Load configuration from YAML file"""
        try:
            if self.config_file.exists():
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    return yaml.safe_load(f)
            else:
                self.app.log_message(f"⚠️ Configuration file not found: {self.config_file}")
                return self._get_default_config()
        except Exception as e:
            self.app.log_message(f"❌ Error loading config: {str(e)}")
            return self._get_default_config()
    
    def _get_default_config(self):
        """Get default configuration"""
        return {
            'departments': ['AT', 'EK', 'EW', 'IS', 'MA', 'PE', 'PN', 'QK', 'QM', 'RA', 'RM', 'TS', 'VB', 'VS'],
            'recipients': {
                'AT': ['mustafa.massli@zyagnum.com'],
                'EK': ['mustafa.massli@zyagnum.com'],
                'EW': ['mustafa.massli@zyagnum.com'],
                'IS': ['mustafa.massli@zyagnum.com'],
                'MA': ['mustafa.massli@zyagnum.com'],
                'PE': ['mustafa.massli@zyagnum.com'],
                'PN': ['mustafa.massli@zyagnum.com'],
                'QK': ['mustafa.massli@zyagnum.com'],
                'QM': ['mustafa.massli@zyagnum.com'],
                'default': ['mustafa.massli@zyagnum.com']
            },
            'email_settings': {
                'cc': ['mustafa.massli@zyagnum.com'],
                'subject_template': 'Aktualitätsprüfung Dokumente Abteilung {department} - {date}'
            },
            'reminder_settings': {
                'check_dates': ['01-08', '07-08'],
                'remind_after_days': 3
            }
        }
    
    def check_and_prompt_halfyear_reminder(self):
        """Check if reminder should be shown and prompt user"""
        try:
            current_date = datetime.now()
            current_mm_dd = current_date.strftime("%m-%d")
            
            # Check if current date matches reminder dates
            if current_mm_dd in self.config['reminder_settings']['check_dates']:
                self.app.log_message(f"📅 Half-year reminder check: {current_mm_dd}")
                
                # Check if already sent this half-year
                halfyear_key = self._get_halfyear_key(current_date)
                
                if not self.has_sent_halfyear(halfyear_key):
                    self.app.log_message(f"📧 Half-year reminder needed for: {halfyear_key}")
                    
                    # Prompt user
                    response = messagebox.askyesno(
                        "Half-Year Deadline Reminder",
                        f"Es ist Zeit für die {halfyear_key} Aktualitätsprüfung.\n\n"
                        "Möchten Sie die Deadline-E-Mails jetzt senden?",
                        icon='question'
                    )
                    
                    if response:
                        self.app.log_message("✅ User chose to send now")
                        self._send_all_department_deadlines(halfyear_key)
                    else:
                        # Ask if they want to be reminded later
                        remind_later = messagebox.askyesno(
                            "Reminder",
                            "Möchten Sie in 3 Tagen erneut erinnert werden?",
                            icon='question'
                        )
                        
                        if remind_later:
                            remind_date = current_date + timedelta(days=3)
                            self.schedule_remind_later(halfyear_key, remind_date)
                            self.app.log_message(f"⏰ Reminder scheduled for {remind_date.strftime('%Y-%m-%d')}")
                        else:
                            self.app.log_message("❌ User declined reminder")
                else:
                    self.app.log_message(f"✅ Already sent for {halfyear_key}")
            else:
                self.app.log_message(f"📅 No reminder needed for {current_mm_dd}")
                
        except Exception as e:
            self.app.log_message(f"❌ Error in half-year reminder check: {str(e)}")
    
    def _get_halfyear_key(self, date):
        """Get half-year key for the given date"""
        year = date.year
        if date.month <= 6:
            return f"H1_{year}"
        else:
            return f"H2_{year}"
    
    def has_sent_halfyear(self, key):
        """Check if already sent for this half-year"""
        try:
            if not self.tracking_file.exists():
                return False
            
            with open(self.tracking_file, 'r', encoding='utf-8') as f:
                tracking_data = json.load(f)
            
            return key in tracking_data and tracking_data[key].get('sent', False)
        except Exception as e:
            self.app.log_message(f"❌ Error checking half-year status: {str(e)}")
            return False
    
    def record_halfyear_sent(self, key):
        """Record that half-year was sent"""
        try:
            tracking_data = {}
            if self.tracking_file.exists():
                with open(self.tracking_file, 'r', encoding='utf-8') as f:
                    tracking_data = json.load(f)
            
            tracking_data[key] = {
                'sent': True,
                'sent_date': datetime.now().isoformat(),
                'remind_date': None
            }
            
            with open(self.tracking_file, 'w', encoding='utf-8') as f:
                json.dump(tracking_data, f, indent=2, ensure_ascii=False)
            
            self.app.log_message(f"✅ Recorded half-year sent: {key}")
        except Exception as e:
            self.app.log_message(f"❌ Error recording half-year sent: {str(e)}")
    
    def schedule_remind_later(self, key, remind_date):
        """Schedule a reminder for later"""
        try:
            tracking_data = {}
            if self.tracking_file.exists():
                with open(self.tracking_file, 'r', encoding='utf-8') as f:
                    tracking_data = json.load(f)
            
            if key not in tracking_data:
                tracking_data[key] = {}
            
            tracking_data[key]['remind_date'] = remind_date.isoformat()
            
            with open(self.tracking_file, 'w', encoding='utf-8') as f:
                json.dump(tracking_data, f, indent=2, ensure_ascii=False)
            
            self.app.log_message(f"⏰ Scheduled reminder for {key} on {remind_date.strftime('%Y-%m-%d')}")
        except Exception as e:
            self.app.log_message(f"❌ Error scheduling reminder: {str(e)}")
    
    def generate_department_deadline_excel(self, department_code, source_file, date_range):
        """Generate Excel file with filtered deadline data for a department"""
        workbook = None
        new_workbook = None
        
        try:
            # Only show detailed logs for QK department
            is_qk = department_code == 'QK'
            
            if is_qk:
                self.app.log_message(f"📊 Checking deadlines for department {department_code}")
            
            # Add file access check
            import os
            if not os.access(source_file, os.R_OK):
                self.app.log_message(f"❌ Cannot access file: {source_file} - please close it if open in Excel")
                return None
            
            # Load Excel file with openpyxl to preserve formatting
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            import warnings
            warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
            
            workbook = load_workbook(source_file, data_only=True)
            
            # Find matching sheet for department
            matching_sheet_name = None
            for sheet_name in workbook.sheetnames:
                if department_code.lower() in sheet_name.lower():
                    matching_sheet_name = sheet_name
                    break
            
            if matching_sheet_name is None:
                # Try to find any sheet that might contain the department
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    # Check if any cell contains the department code
                    for row in worksheet.iter_rows():
                        for cell in row:
                            if cell.value and department_code.lower() in str(cell.value).lower():
                                matching_sheet_name = sheet_name
                                break
                        if matching_sheet_name:
                            break
                    if matching_sheet_name:
                        break
            
            if matching_sheet_name is None:
                if is_qk:
                    self.app.log_message(f"⚠️ No matching sheet found for department {department_code}")
                workbook.close()
                return None
            
            worksheet = workbook[matching_sheet_name]
            
            # Filter by deadline column (column H = index 8)
            deadline_col = 8  # Column H
            
            # Get date range
            start_date, end_date = date_range
            
            # TEMPORARY DEBUG: Uncomment the next 2 lines to test with hardcoded range
            # start_date = datetime(2025, 11, 1)
            # end_date = datetime(2025, 11, 30)
            
            if is_qk:
                self.app.log_message(f"📆 Filtering deadlines between {start_date} and {end_date}")
            
            # Find rows with "A" in column A and check dates in column H
            matching_rows = []
            
            for row_num in range(10, worksheet.max_row + 1):  # Start from row 10 where data begins
                col_a_cell = worksheet.cell(row=row_num, column=1)
                if not col_a_cell.value or str(col_a_cell.value).strip() != "A":
                    continue
                
                deadline_cell = worksheet.cell(row=row_num, column=deadline_col)
                if not deadline_cell.value:
                    continue
                    
                try:
                    deadline_date = pd.to_datetime(deadline_cell.value, errors='coerce')
                    if pd.isna(deadline_date):
                        continue
                    
                    if start_date <= deadline_date <= end_date:
                        matching_rows.append(row_num)
                        if is_qk:
                            self.app.log_message(f"✅ Found matching deadline: {deadline_date} in row {row_num}")
                except Exception as e:
                    if is_qk:
                        self.app.log_message(f"⚠️ Error processing row {row_num}: {e}")
                    continue
            
            if len(matching_rows) == 0:
                if is_qk:
                    self.app.log_message(f"⚠️ No deadlines found for department {department_code}")
                workbook.close()
                return None
            
            # Create new workbook with all formatting preserved
            from openpyxl import Workbook
            new_workbook = Workbook()
            new_worksheet = new_workbook.active
            new_worksheet.title = f"{department_code}_Fristen"
            
            # Copy header row (row 9) with formatting
            if is_qk:
                self.app.log_message(f"📝 Copying header row 9")
            
            # Copy column widths first
            for col in range(1, 9):  # Columns A to H
                col_letter = get_column_letter(col)
                new_worksheet.column_dimensions[col_letter].width = worksheet.column_dimensions[col_letter].width
            
            # Copy header row (row 9) with all formatting
            for col in range(1, 9):
                try:
                    source_cell = worksheet.cell(row=9, column=col)
                    target_cell = new_worksheet.cell(row=1, column=col)
                    
                    # Copy value
                    target_cell.value = source_cell.value
                    
                    # Copy all formatting properties
                    if source_cell.has_style:
                        target_cell.font = Font(
                            name=source_cell.font.name,
                            size=source_cell.font.size,
                            bold=source_cell.font.bold,
                            italic=source_cell.font.italic,
                            color=source_cell.font.color
                        )
                        
                        target_cell.border = Border(
                            left=Side(border_style=source_cell.border.left.style,
                                     color=source_cell.border.left.color),
                            right=Side(border_style=source_cell.border.right.style,
                                      color=source_cell.border.right.color),
                            top=Side(border_style=source_cell.border.top.style,
                                    color=source_cell.border.top.color),
                            bottom=Side(border_style=source_cell.border.bottom.style,
                                       color=source_cell.border.bottom.color)
                        )
                        
                        target_cell.fill = PatternFill(
                            fill_type=source_cell.fill.fill_type,
                            start_color=source_cell.fill.start_color,
                            end_color=source_cell.fill.end_color
                        )
                        
                        target_cell.number_format = source_cell.number_format
                        
                        target_cell.alignment = Alignment(
                            horizontal=source_cell.alignment.horizontal,
                            vertical=source_cell.alignment.vertical,
                            wrap_text=source_cell.alignment.wrap_text,
                            shrink_to_fit=source_cell.alignment.shrink_to_fit,
                            indent=source_cell.alignment.indent
                        )
                except Exception as e:
                    if is_qk:
                        self.app.log_message(f"⚠️ Error copying header cell at col {col}: {e}")
            
            # Copy matching rows with all formatting
            for new_row_idx, original_row in enumerate(matching_rows, start=2):
                try:
                    for col in range(1, 9):
                        try:
                            source_cell = worksheet.cell(row=original_row, column=col)
                            target_cell = new_worksheet.cell(row=new_row_idx, column=col)
                            
                            # Copy value with proper date formatting
                            if isinstance(source_cell.value, datetime):
                                target_cell.value = source_cell.value
                                # Use the original source cell's number format to preserve date formatting
                                target_cell.number_format = source_cell.number_format
                            else:
                                target_cell.value = source_cell.value
                            
                            # Copy all formatting properties
                            if source_cell.has_style:
                                target_cell.font = source_cell.font.copy()
                                target_cell.border = source_cell.border.copy()
                                target_cell.fill = source_cell.fill.copy()
                                target_cell.number_format = source_cell.number_format
                                target_cell.alignment = source_cell.alignment.copy()
                        except Exception as e:
                            if is_qk:
                                self.app.log_message(f"⚠️ Error copying cell at row {original_row}, col {col}: {e}")
                except Exception as e:
                    if is_qk:
                        self.app.log_message(f"⚠️ Error copying row {original_row}: {e}")
            
            # Generate output filename - one file per department with all deadlines
            current_date = datetime.now().strftime("%Y%m%d")
            output_filename = f"{department_code}_Fristen_{current_date}.xlsx"
            output_path = Path.home() / "Documents" / output_filename
            
            # Handle filename collisions
            counter = 1
            original_path = output_path
            while output_path.exists():
                stem = original_path.stem
                suffix = original_path.suffix
                output_path = original_path.parent / f"{stem}_{counter}{suffix}"
                counter += 1
            
            # Save with all formatting
            new_workbook.save(str(output_path))
            
            self.app.log_message(f"✅ Generated department Excel with {len(matching_rows)} deadlines: {output_path.name}")
            return output_path
            
        except Exception as e:
            self.app.log_message(f"❌ Error generating deadline Excel: {str(e)}")
            return None
            
        finally:
            # Ensure proper cleanup
            try:
                if workbook is not None:
                    workbook.close()
            except Exception as e:
                pass
            try:
                if new_workbook is not None:
                    new_workbook.close()
            except Exception as e:
                pass
    
    def send_deadline_email(self, recipient_emails, attachment, department_code, date_range):
        """Send deadline email with attachment"""
        try:
            self.app.log_message(f"📧 Sending deadline email for department {department_code}")
            
            # Create email content
            subject, body = self._create_email_content(department_code, date_range)
            
            # Send email
            if platform.system() == "Windows":
                success = self._send_email_windows(recipient_emails, subject, body, attachment)
            else:
                success = self._send_email_smtp(recipient_emails, subject, body, attachment)
            
            if success:
                self.app.log_message(f"✅ Deadline email sent successfully for {department_code}")
            else:
                self.app.log_message(f"❌ Failed to send deadline email for {department_code}")
            
            return success
            
        except Exception as e:
            self.app.log_message(f"❌ Error sending deadline email: {str(e)}")
            return False
    
    def _create_email_content(self, department_code, date_range):
        """Create email subject and body"""
        current_date = datetime.now().strftime("%d.%m.%Y")
        start_date, end_date = date_range
        
        # Determine half-year and year
        current_date_obj = datetime.now()
        if current_date_obj.month <= 6:
            halbjahr = 1
        else:
            halbjahr = 2
        
        # Calculate deadline date (March 30 for H1, September 30 for H2)
        if halbjahr == 1:
            deadline_date = datetime(current_date_obj.year, 3, 30)
        else:
            deadline_date = datetime(current_date_obj.year, 9, 30)
        
        subject = self.config['email_settings']['subject_template'].format(
            department=department_code,
            date=current_date
        )
        
        body = f"""Liebe Kolleg:innen,

Es ist wieder eine Aktualitätsprüfung von einigen Dokumenten erforderlich.

Der Workflow ist wie gehabt:

1. Ich überprüfe am Anfang jedes Halbjahrs die Einträge in unserer Dokumentenübersichtsliste (QM-LIS-001_Liste der Dokumente) und informiere euch per E-Mail über Dokumente, die auf Aktualität und Korrektheit überprüft werden müssen.
2. Die zuständige Person (Abteilungsleiter/Prozesseigner) prüft den Inhalt sowie Form und Struktur der Dokumente.
3. Die zuständige Person (Abteilungsleiter/Prozesseigner) meldet die Überprüfung via E-Mail an mich mit folgenden Informationen:
   • Welches Dokument von euch überprüft wurde
   • Wann die Prüfung erfolgte
   • Ob das Dokument aktuell und korrekt ist
   • Ob eine Aktualisierung des Dokuments erforderlich ist und ob ihr ein Meeting benötigt
   • Sonstige Informationen

4. Ich trage die Informationen in die QM-LIS-001_Liste der Dokumente ein. Ich bitte euch, selbst keine Einträge oder Änderungen in der Liste vorzunehmen!

Folgend findet ihr im Anhang mit Dokumenten, welche im {halbjahr}. Halbjahr {current_date_obj.year} überprüft werden müssen:

Wir bitten euch, die Dokumente bis zum {deadline_date.strftime('%d.%m.%Y')} auf Aktualität und Korrektheit zu überprüfen.

Vielen Dank für eure Mitarbeit!

Falls ihr Fragen dazu habt, lasst es mich gerne wissen.

Liebe Grüße,
Sarah und Mustafa"""
        
        return subject, body
    
    def _send_email_windows(self, recipient_emails, subject, body, attachment):
        """Send email using Windows COM (Outlook)"""
        try:
            import win32com.client
            
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)  # 0 = olMailItem
            
            # Set recipients
            mail.To = "; ".join(recipient_emails)
            mail.CC = "; ".join(self.config['email_settings']['cc'])
            mail.Subject = subject
            mail.Body = body
            
            # Add attachment
            if attachment and Path(attachment).exists():
                mail.Attachments.Add(str(attachment))
            
            # Send
            mail.Send()
            return True
            
        except ImportError:
            self.app.log_message("⚠️ win32com not available, falling back to SMTP")
            return self._send_email_smtp(recipient_emails, subject, body, attachment)
        except Exception as e:
            self.app.log_message(f"❌ Error sending email via Windows COM: {str(e)}")
            return False
    
    def _send_email_smtp(self, recipient_emails, subject, body, attachment):
        """Send email using SMTP (fallback)"""
        try:
            # This is a placeholder - you would need to configure SMTP settings
            self.app.log_message("⚠️ SMTP email sending not configured")
            return False
        except Exception as e:
            self.app.log_message(f"❌ Error sending email via SMTP: {str(e)}")
            return False
    
    def _send_all_department_deadlines(self, halfyear_key):
        """Send deadline emails for all departments"""
        try:
            current_date = datetime.now()
            
            # Determine date range based on half-year
            if "H1" in halfyear_key:
                start_date = datetime(current_date.year, 1, 1)
                end_date = datetime(current_date.year, 6, 30)
            else:
                start_date = datetime(current_date.year, 7, 1)
                end_date = datetime(current_date.year, 12, 31)
            
            date_range = (start_date, end_date)
            
            # Get source Excel file from GUI
            source_file = self.app.excel_entry.get().strip()
            if not source_file or not Path(source_file).exists():
                messagebox.showerror("Error", "Please select a valid Excel file first")
                return
            
            success_count = 0
            
            for department in self.config['departments']:
                try:
                    self.app.log_message(f"📊 Processing department: {department}")
                    
                    # Generate Excel file
                    excel_path = self.generate_department_deadline_excel(
                        department, source_file, date_range
                    )
                    
                    # Only send email if Excel file was generated (has deadlines)
                    if excel_path is not None:
                        # Get recipients
                        recipients = self.config['recipients'].get(
                            department, 
                            self.config['recipients']['default']
                        )
                        
                        # Send email
                        if self.send_deadline_email(recipients, excel_path, department, date_range):
                            success_count += 1
                    else:
                        self.app.log_message(f"ℹ️ Skipping email for {department} - no deadlines found")
                    
                except Exception as e:
                    self.app.log_message(f"❌ Error processing department {department}: {str(e)}")
            
            # Record success
            if success_count > 0:
                self.record_halfyear_sent(halfyear_key)
                messagebox.showinfo(
                    "Success", 
                    f"Successfully sent {success_count} deadline emails for {halfyear_key}"
                )
            else:
                messagebox.showerror(
                    "Error", 
                    "Failed to send any deadline emails. Please check the logs."
                )
                
        except Exception as e:
            self.app.log_message(f"❌ Error sending all department deadlines: {str(e)}")
            messagebox.showerror("Error", f"Error sending deadlines: {str(e)}")
    
    def reset_halfyear_status(self):
        """Reset half-year tracking status"""
        try:
            if self.tracking_file.exists():
                self.tracking_file.unlink()
                self.app.log_message("✅ Half-year tracking status reset")
            else:
                self.app.log_message("ℹ️ No tracking file found to reset")
        except Exception as e:
            self.app.log_message(f"❌ Error resetting half-year status: {str(e)}")
    
    def show_tracking_status(self):
        """Show current tracking status and provide options to generate deadline Excel files"""
        try:
            current_date = datetime.now()
            current_halfyear = self._get_halfyear_key(current_date)
            
            # Get tracking data if it exists
            tracking_data = {}
            if self.tracking_file.exists():
                with open(self.tracking_file, 'r', encoding='utf-8') as f:
                    tracking_data = json.load(f)
            
            # Build status text
            status_text = f"Deadline Tracking Status - {current_date.strftime('%d.%m.%Y')}\n"
            status_text += f"Current Half-Year: {current_halfyear}\n\n"
            
            # Show tracking data
            if tracking_data:
                status_text += "Tracking History:\n"
                for key, data in tracking_data.items():
                    status = "✅ Sent" if data.get('sent', False) else "❌ Not sent"
                    sent_date = data.get('sent_date', 'N/A')
                    remind_date = data.get('remind_date', 'N/A')
                    
                    status_text += f"  {key}: {status}\n"
                    if sent_date != 'N/A':
                        status_text += f"    Sent: {sent_date}\n"
                    if remind_date != 'N/A':
                        status_text += f"    Remind: {remind_date}\n"
                    status_text += "\n"
            else:
                status_text += "No tracking data found.\n\n"
            
            # Show current status
            current_sent = tracking_data.get(current_halfyear, {}).get('sent', False)
            status_text += f"Current Status:\n"
            status_text += f"  {current_halfyear}: {'✅ Sent' if current_sent else '❌ Not sent'}\n\n"
            
            # Add action options
            status_text += "Available Actions:\n"
            status_text += "• Generate deadline Excel files and send emails\n"
            status_text += "• Send deadline emails to all departments\n"
            status_text += "• Reset tracking status\n\n"
            
            # Show dialog with options
            response = messagebox.askyesnocancel(
                "Deadline Tracking Status",
                status_text + "\nWould you like to generate deadline Excel files now?",
                icon='question'
            )
            
            if response is True:  # Yes - Generate Excel files
                self._generate_current_deadline_excel_files()
            elif response is False:  # No - Show more options
                self._show_additional_options()
            
        except Exception as e:
            self.app.log_message(f"❌ Error showing tracking status: {str(e)}")
            messagebox.showerror("Error", f"Error showing tracking status: {str(e)}")
    
    def _generate_current_deadline_excel_files(self):
        """Generate deadline Excel files for the current half-year and optionally send emails"""
        try:
            # Get source Excel file from GUI
            source_file = self.app.excel_entry.get().strip()
            if not source_file or not Path(source_file).exists():
                messagebox.showerror("Error", "Please select a valid Excel file first")
                return
            
            current_date = datetime.now()
            current_halfyear = self._get_halfyear_key(current_date)
            
            # Determine date range
            if "H1" in current_halfyear:
                start_date = datetime(current_date.year, 1, 1)
                end_date = datetime(current_date.year, 6, 30)
            else:
                start_date = datetime(current_date.year, 7, 1)
                end_date = datetime(current_date.year, 12, 31)
            
            date_range = (start_date, end_date)
            
            # Generate Excel files for all departments
            generated_files = []
            departments_with_deadlines = []
            departments_without_deadlines = []
            
            for department in self.config['departments']:
                try:
                    excel_path = self.generate_department_deadline_excel(
                        department, source_file, date_range
                    )
                    if excel_path is not None:
                        generated_files.append((department, excel_path))
                        departments_with_deadlines.append(department)
                        self.app.log_message(f"✅ Generated: {excel_path.name}")
                    else:
                        departments_without_deadlines.append(department)
                        self.app.log_message(f"ℹ️ No deadlines found for {department}")
                except Exception as e:
                    self.app.log_message(f"❌ Error generating Excel for {department}: {str(e)}")
            
            # Show results and ask if user wants to send emails
            if generated_files:
                result_text = f"Generated {len(generated_files)} deadline Excel files:\n\n"
                for dept, file_path in generated_files:
                    result_text += f"• {dept}: {file_path.name}\n"
                result_text += f"\nFiles saved to: {Path.home() / 'Documents'}"
                
                if departments_without_deadlines:
                    result_text += f"\n\nDepartments with no deadlines: {', '.join(departments_without_deadlines)}"
                
                result_text += "\n\nWould you like to send deadline emails to departments with deadlines?"
                
                send_emails = messagebox.askyesno(
                    "Excel Generation Complete", 
                    result_text,
                    icon='question'
                )
                
                if send_emails:
                    self._send_deadline_emails_with_files(generated_files, date_range, current_halfyear)
            else:
                messagebox.showerror("Error", "No Excel files were generated. Check the logs for details.")
                
        except Exception as e:
            self.app.log_message(f"❌ Error generating deadline Excel files: {str(e)}")
            messagebox.showerror("Error", f"Error generating Excel files: {str(e)}")
    
    def _send_deadline_emails_with_files(self, generated_files, date_range, halfyear_key):
        """Send deadline emails with the generated Excel files"""
        try:
            self.app.log_message(f"📧 Starting to send deadline emails for {len(generated_files)} departments")
            
            success_count = 0
            failed_departments = []
            
            for department, excel_path in generated_files:
                try:
                    # Get recipients for this department
                    recipients = self.config['recipients'].get(
                        department, 
                        self.config['recipients']['default']
                    )
                    
                    self.app.log_message(f"📧 Sending email for department {department} to {recipients}")
                    
                    # Send email with attachment
                    if self.send_deadline_email(recipients, excel_path, department, date_range):
                        success_count += 1
                        self.app.log_message(f"✅ Email sent successfully for {department}")
                    else:
                        failed_departments.append(department)
                        self.app.log_message(f"❌ Failed to send email for {department}")
                        
                except Exception as e:
                    failed_departments.append(department)
                    self.app.log_message(f"❌ Error sending email for {department}: {str(e)}")
            
            # Show results
            if success_count > 0:
                # Record that emails were sent
                self.record_halfyear_sent(halfyear_key)
                
                result_text = f"Successfully sent {success_count} deadline emails!\n\n"
                result_text += f"Sent to departments: {', '.join([dept for dept, _ in generated_files if dept not in failed_departments])}"
                
                if failed_departments:
                    result_text += f"\n\nFailed departments: {', '.join(failed_departments)}"
                
                messagebox.showinfo("Email Sending Complete", result_text)
            else:
                messagebox.showerror(
                    "Email Sending Failed", 
                    "Failed to send any deadline emails. Please check the logs for details."
                )
                
        except Exception as e:
            self.app.log_message(f"❌ Error sending deadline emails: {str(e)}")
            messagebox.showerror("Error", f"Error sending deadline emails: {str(e)}")
    
    def _show_additional_options(self):
        """Show additional options for deadline management"""
        options = [
            ("Send Deadline Emails", "Send deadline emails to all departments"),
            ("Reset Tracking", "Clear all tracking data"),
            ("Cancel", "Close without action")
        ]
        
        # Create custom dialog
        dialog = tk.Toplevel()
        dialog.title("Deadline Management Options")
        dialog.geometry("400x300")
        dialog.transient(self.app.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (400 // 2)
        y = (dialog.winfo_screenheight() // 2) - (300 // 2)
        dialog.geometry(f"400x300+{x}+{y}")
        
        # Add content
        ttk.Label(dialog, text="Select an action:", font=("Arial", 12, "bold")).pack(pady=10)
        
        for text, description in options:
            frame = ttk.Frame(dialog)
            frame.pack(fill=tk.X, padx=20, pady=5)
            
            btn = ttk.Button(
                frame, 
                text=text,
                command=lambda t=text: self._handle_option_selection(t, dialog)
            )
            btn.pack(fill=tk.X)
            
            ttk.Label(frame, text=description, font=("Arial", 9)).pack(anchor=tk.W, padx=5)
        
        # Wait for dialog to close
        dialog.wait_window()
    
    def _handle_option_selection(self, option, dialog):
        """Handle option selection from the dialog"""
        dialog.destroy()
        
        if option == "Send Deadline Emails":
            # First generate Excel files, then send emails
            self._generate_and_send_deadline_emails()
        elif option == "Reset Tracking":
            self.reset_halfyear_status()
            messagebox.showinfo("Reset Complete", "Deadline tracking status has been reset.")
        # Cancel does nothing (dialog already closed)
    
    def _generate_and_send_deadline_emails(self):
        """Generate Excel files and send emails in one workflow"""
        try:
            # Get source Excel file from GUI
            source_file = self.app.excel_entry.get().strip()
            if not source_file or not Path(source_file).exists():
                messagebox.showerror("Error", "Please select a valid Excel file first")
                return
            
            current_date = datetime.now()
            current_halfyear = self._get_halfyear_key(current_date)
            
            # Determine date range
            if "H1" in current_halfyear:
                start_date = datetime(current_date.year, 1, 1)
                end_date = datetime(current_date.year, 6, 30)
            else:
                start_date = datetime(current_date.year, 7, 1)
                end_date = datetime(current_date.year, 12, 31)
            
            date_range = (start_date, end_date)
            
            # Generate Excel files for all departments
            generated_files = []
            departments_with_deadlines = []
            departments_without_deadlines = []
            
            for department in self.config['departments']:
                try:
                    excel_path = self.generate_department_deadline_excel(
                        department, source_file, date_range
                    )
                    if excel_path is not None:
                        generated_files.append((department, excel_path))
                        departments_with_deadlines.append(department)
                        self.app.log_message(f"✅ Generated: {excel_path.name}")
                    else:
                        departments_without_deadlines.append(department)
                        self.app.log_message(f"ℹ️ No deadlines found for {department}")
                except Exception as e:
                    self.app.log_message(f"❌ Error generating Excel for {department}: {str(e)}")
            
            # Send emails with the generated files
            if generated_files:
                self.app.log_message(f"📧 Sending emails to {len(departments_with_deadlines)} departments with deadlines")
                self._send_deadline_emails_with_files(generated_files, date_range, current_halfyear)
            else:
                messagebox.showinfo("No Deadlines Found", 
                    f"No deadlines found for any departments in the current half-year ({current_halfyear}).\n\n"
                    f"Checked departments: {', '.join(self.config['departments'])}")
                
        except Exception as e:
            self.app.log_message(f"❌ Error in generate and send workflow: {str(e)}")
            messagebox.showerror("Error", f"Error in generate and send workflow: {str(e)}") 