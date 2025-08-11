import os
import sys
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment
# Import Font with alias to avoid conflict with Spire.XLS
from openpyxl.styles import Font as OpenpyxlFont
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.drawing.image import Image
import warnings

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import re
import tempfile
import subprocess
import platform
from gui.dialogs import ExcelCellInputDialog

# Try to import Spire.XLS for advanced watermarking
try:
    from spire.xls import *
    from spire.xls.common import *
    from System.Drawing import Color  # Correct color import
    SPIRE_XLS_AVAILABLE = True
except ImportError:
    SPIRE_XLS_AVAILABLE = False
except Exception:
    # Handle other initialization errors
    SPIRE_XLS_AVAILABLE = False

# Try to import win32com for Excel automation
try:
    import win32com.client
    WIN32COM_AVAILABLE = True
except ImportError:
    WIN32COM_AVAILABLE = False

# Suppress openpyxl warnings about print areas
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', message='Print area cannot be set to Defined name')


class ExcelOperations:
    """Handles Excel file operations using pandas and openpyxl"""

    def __init__(self, app):
        self.app = app

    def is_excel_file(self, file_path):
        """Check if file is an Excel file"""
        excel_extensions = ['.xlsx', '.xls', '.xlsm', '.xlsb']
        return Path(file_path).suffix.lower() in excel_extensions

    def browse_excel_file(self, entry_widget):
        """Browse for an Excel file and update the entry widget"""
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[
                ("Excel files", "*.xlsx *.xls *.xlsm *.xlsb"),
                ("Excel 2007+", "*.xlsx"),
                ("Excel 97-2003", "*.xls"),
                ("Excel Macro-enabled", "*.xlsm"),
                ("Excel Binary", "*.xlsb"),
                ("All files", "*.*")
            ]
        )
        
        if file_path:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, file_path)
            self.app.log_message(f"📊 Excel file updated and saved: {Path(file_path).name}")
            self.app.config_manager.save_config()  # Auto-save persistent file
            return file_path
        return None

    def read_excel_file(self, file_path, sheet_name=None):
        """Read Excel file using pandas"""
        try:
            if sheet_name is None:
                # Read all sheets
                excel_file = pd.ExcelFile(file_path)
                sheets = {}
                for sheet in excel_file.sheet_names:
                    sheets[sheet] = pd.read_excel(file_path, sheet_name=sheet)
                return sheets
            else:
                # Read specific sheet
                return pd.read_excel(file_path, sheet_name=sheet_name)
        except Exception as e:
            self.app.log_message(f"❌ Error reading Excel file: {str(e)}")
            raise

    def write_excel_file(self, data, file_path, sheet_name='Sheet1'):
        """Write data to Excel file using pandas"""
        try:
            if isinstance(data, dict):
                # Multiple sheets
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    for sheet_name, df in data.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                # Single sheet
                data.to_excel(file_path, sheet_name=sheet_name, index=False)
            
            self.app.log_message(f"✅ Excel file saved: {Path(file_path).name}")
            return True
        except Exception as e:
            self.app.log_message(f"❌ Error writing Excel file: {str(e)}")
            raise

    def get_excel_info(self, file_path):
        """Get information about Excel file (sheets, dimensions, etc.)"""
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                workbook = load_workbook(file_path, read_only=True)
                
            info = {
                'sheets': workbook.sheetnames,
                'sheet_count': len(workbook.sheetnames),
                'file_size': Path(file_path).stat().st_size,
                'file_path': file_path
            }
            
            # Get dimensions for each sheet
            sheet_info = {}
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                sheet_info[sheet_name] = {
                    'max_row': worksheet.max_row,
                    'max_column': worksheet.max_column,
                    'dimensions': f"{worksheet.max_row} rows × {worksheet.max_column} columns"
                }
            
            info['sheet_details'] = sheet_info
            workbook.close()
            return info
        except Exception as e:
            self.app.log_message(f"❌ Error getting Excel info: {str(e)}")
            raise

    def validate_excel_file(self, file_path):
        """Validate that file is a readable Excel file"""
        try:
            if not self.is_excel_file(file_path):
                return False, "File is not an Excel file"
            
            if not Path(file_path).exists():
                return False, "File does not exist"
            
            # Try to open the file
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                workbook = load_workbook(file_path, read_only=True)
            workbook.close()
            return True, "Valid Excel file"
        except Exception as e:
            return False, f"Error reading Excel file: {str(e)}"

    def create_excel_summary(self, file_path):
        """Create a summary of Excel file contents"""
        try:
            info = self.get_excel_info(file_path)
            summary = f"Excel File Summary: {Path(file_path).name}\n"
            summary += f"File size: {self.format_file_size(info['file_size'])}\n"
            summary += f"Number of sheets: {info['sheet_count']}\n\n"
            
            summary += "Sheets:\n"
            for sheet_name, details in info['sheet_details'].items():
                summary += f"  - {sheet_name}: {details['dimensions']}\n"
            
            return summary
        except Exception as e:
            self.app.log_message(f"❌ Error creating Excel summary: {str(e)}")
            raise

    def format_file_size(self, size_bytes):
        """Format file size in human readable format"""
        if size_bytes == 0:
            return "0 B"
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size_bytes < 1024.0:
                return f"{size_bytes:.1f} {unit}"
            size_bytes /= 1024.0
        return f"{size_bytes:.1f} TB"

    def export_to_excel(self, data, file_path, sheet_name='Sheet1'):
        """Export data to Excel file"""
        try:
            if isinstance(data, pd.DataFrame):
                data.to_excel(file_path, sheet_name=sheet_name, index=False)
            elif isinstance(data, dict):
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    for sheet_name, df in data.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                raise ValueError("Data must be a pandas DataFrame or dict of DataFrames")
            
            self.app.log_message(f"✅ Data exported to Excel: {Path(file_path).name}")
            return True
        except Exception as e:
            self.app.log_message(f"❌ Error exporting to Excel: {str(e)}")
            raise



    def update_excel_tracking(self, doc_prefix, attachment_filename, has_multiple_formats=False, all_files_in_group=None):
        """
        Update Excel tracking file with file replacement information.
        
        Args:
            doc_prefix (str): The document prefix to match in column B
            attachment_filename (str): The attachment filename to split and insert
            has_multiple_formats (bool): Whether this file has multiple format versions
            all_files_in_group (list): List of all files in the same group (for multiple formats)
        """
        try:
            # Normalize the doc_prefix by removing trailing hyphens to fix matching issues
            doc_prefix = (doc_prefix or "").rstrip("-")
            
            self.app.log_message(f"🔍 Starting Excel tracking update...")
            
            # Store current document information for dialogs
            self.current_doc_prefix = doc_prefix
            self.current_attachment_filename = attachment_filename
            
            # Get Excel file path from app's Excel input field
            excel_file_path = self.app.excel_entry.get().strip()
            
            if not excel_file_path or not os.path.exists(excel_file_path):
                self.app.log_message("❌ Excel tracking file not found or not specified")
                return False
            
            # Get target directory from GUI input
            target_dir = self.app.target_entry.get().strip()
            
            if not target_dir:
                self.app.log_message("❌ Target directory not specified")
                return False
            
            # Load the Excel workbook
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                workbook = load_workbook(excel_file_path)
            
            match_found = False
            old_status = None
            matched_sheet = None
            last_matched_row = None
            all_matches = []  # Store all matches to find the last one
            
            # Loop through all sheets to find all matches
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                self.app.log_message(f"🔍 Searching sheet '{sheet_name}' for prefix '{doc_prefix}'")
                # Check each row starting from row 9
                for row_num in range(9, worksheet.max_row + 1):
                    # Check column B (index 1) for prefix match
                    cell_value = worksheet.cell(row=row_num, column=2).value
                    if cell_value:
                        self.app.log_message(f"🔍 Row {row_num}, Column B: '{cell_value}' vs prefix '{doc_prefix}'")
                        if self.has_same_document_id(attachment_filename, cell_value):
                            self.app.log_message(f"✅ MATCH FOUND at row {row_num}: '{cell_value}'")
                        # --- BEGIN: In-place update for 'C' rows with exact document ID match ---
                        # Check for 'C' in column A and exact document ID match in column B
                        col_a_value = worksheet.cell(row=row_num, column=1).value
                        
                        # Get Document ID from attachment filename
                        doc_id_from_filename = self.extract_document_id(attachment_filename)
                        
                        if (
                            col_a_value == 'C' and  # Status is 'C'
                            doc_id_from_filename and  # Valid Document ID extracted
                            str(cell_value).strip() == doc_id_from_filename  # Exact match with Column B
                        ):
                            self.app.log_message(f"🟢 In-place update: Found 'C' row with exact document ID match '{doc_id_from_filename}' at row {row_num} in sheet '{sheet_name}'")
                            # Set column A to 'A'
                            worksheet.cell(row=row_num, column=1, value='A')
                            # Update file info: split filename parts into columns B, C, D using regex
                            col_b_value, col_c_value, col_d_value = self._split_filename_with_regex(attachment_filename)
                            worksheet.cell(row=row_num, column=2, value=col_b_value)
                            worksheet.cell(row=row_num, column=3, value=col_c_value)
                            worksheet.cell(row=row_num, column=4, value=col_d_value)
                            # Prepare new_row_data for dialog (show only 'New Row' tab)
                            new_row_data = {
                                'E': worksheet.cell(row=row_num, column=5).value or '',
                                'F': worksheet.cell(row=row_num, column=6).value or '',
                                'G': worksheet.cell(row=row_num, column=7).value or ''
                            }
                            document_info = {
                                'filename': getattr(self, 'current_attachment_filename', 'Unknown'),
                                'doc_prefix': getattr(self, 'current_doc_prefix', 'Unknown'),
                                'sheet_name': worksheet.title,
                                'row_num': row_num,
                                'is_v1_file': False
                            }
                            # Show dialog for new row only
                            from gui.dialogs import ExcelCellInputDialog
                            try:
                                self.app.log_message(f"🔍 Creating Excel cell input dialog for: {attachment_filename}")
                                dialog = ExcelCellInputDialog(self.app.root, None, new_row_data, document_info)
                                # Ensure dialog is properly shown
                                self.app.log_message(f"🔍 Showing Excel cell input dialog...")
                                dialog.show_dialog()
                                self.app.root.wait_window(dialog.dialog)
                                self.app.log_message(f"🔍 Excel cell input dialog closed")
                            except Exception as e:
                                self.app.log_message(f"❌ Error creating Excel dialog: {str(e)}")
                                # Fallback: use default values
                                dialog = None
                                dialog.result = {'new_row': {'E': '', 'F': 'aktuell gültig', 'G': '-'}}
                            
                            # Apply user input if confirmed
                            if dialog and dialog.result and 'new_row' in dialog.result:
                                for col, value in dialog.result['new_row'].items():
                                    col_index = {'E': 5, 'F': 6, 'G': 7}[col]
                                    parsed_value = self._parse_date_value(value)
                                    worksheet.cell(row=row_num, column=col_index, value=parsed_value)
                            
                            # Add hyperlink in column J pointing to the file's new location
                            try:
                                target_dir = self.app.target_entry.get().strip()
                                if target_dir:
                                    target_path = Path(target_dir) / attachment_filename
                                    cell_j = worksheet.cell(row=row_num, column=10)  # Column J is index 10
                                    
                                    if target_path.exists():
                                        target_hyperlink = f"file:///{target_path.absolute().as_posix()}"
                                        cell_j.value = str(target_path)
                                        cell_j.hyperlink = target_hyperlink
                                        self.app.log_message(f"🔗 Added hyperlink to column J: {target_path}")
                                    else:
                                        self.app.log_message(f"⚠️ Target document not found at: {target_path}")
                                        cell_j.value = f"{target_path} (Not Found)"
                            except Exception as e:
                                self.app.log_message(f"❌ Error adding hyperlink to column J: {str(e)}")
                            
                            # Save workbook and exit
                            workbook.save(excel_file_path)
                            self.app.log_message(f"✅ In-place update completed and saved for row {row_num} in sheet '{sheet_name}'")
                            return True
                        # --- END: In-place update for 'C' rows with prefix match ---
                        if self.has_same_document_id(attachment_filename, cell_value):
                            # Store match information
                            all_matches.append({
                                'sheet_name': sheet_name,
                                'row_num': row_num,
                                'old_status': worksheet.cell(row=row_num, column=1).value
                            })
                            match_found = True
                            matched_sheet = sheet_name
                            last_matched_row = row_num
            
            # Check if this is a V1.0 file (character beside "V" is "1")
            is_v1_file = False
            v_index = attachment_filename.find('V')
            if v_index != -1 and v_index + 1 < len(attachment_filename):
                char_beside_v = attachment_filename[v_index + 1]
                if char_beside_v == '1':
                    is_v1_file = True
                    self.app.log_message(f"🔍 Detected V1.0 file in Excel tracking: {attachment_filename}")

            # Handle V1.0 files first - they use special document ID matching logic
            if is_v1_file:
                self.app.log_message(f"🚀 V1.0 file detected - using document ID pattern matching (bypassing prefix search)")
                
                # Extract document ID from filename
                doc_id_from_filename = self.extract_document_id(attachment_filename)
                
                # Find the correct sheet and optimal insertion position
                insertion_row = 8  # Default to before row 9
                found_optimal_position = False
                target_sheet = None
                
                # Search through all sheets to find the correct one
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    
                    # First pass: Find the entry with the largest numeric suffix in column B
                    max_suffix = 0
                    max_suffix_row = None
                    
                    for row_num in range(9, worksheet.max_row + 1):
                        col_b_value = str(worksheet.cell(row=row_num, column=2).value or "")
                        
                        if self.has_same_document_id(attachment_filename, col_b_value):
                            try:
                                current_suffix = int(col_b_value[-3:]) if len(col_b_value) >= 3 else 0
                                if current_suffix > max_suffix:
                                    max_suffix = current_suffix
                                    max_suffix_row = row_num
                            except ValueError:
                                continue
                    
                    # If we found entries with the same prefix, check version numbers in column C
                    if max_suffix_row is not None:
                        target_sheet = sheet_name
                        insertion_row = max_suffix_row
                        
                        # Get the version number from the new file
                        new_version = 0
                        v_index = attachment_filename.find('V')
                        if v_index != -1 and v_index + 1 < len(attachment_filename):
                            try:
                                new_version = int(attachment_filename[v_index + 1])
                            except ValueError:
                                pass
                        
                        # Check if there are multiple entries with the same suffix but different versions
                        max_version = 0
                        max_version_row = max_suffix_row
                        
                        for row_num in range(9, worksheet.max_row + 1):
                            col_b_value = str(worksheet.cell(row=row_num, column=2).value or "")
                            col_c_value = str(worksheet.cell(row=row_num, column=3).value or "")
                            
                            if self.has_same_document_id(attachment_filename, col_b_value):
                                try:
                                    current_suffix = int(col_b_value[-3:]) if len(col_b_value) >= 3 else 0
                                    if current_suffix == max_suffix:
                                        # Same suffix, check version in column C
                                        v_match = re.search(r'V(\d+)', col_c_value)
                                        if v_match:
                                            current_version = int(v_match.group(1))
                                            if current_version > max_version:
                                                max_version = current_version
                                                max_version_row = row_num
                                except ValueError:
                                    continue
                        
                        # Use the row with the largest version number
                        insertion_row = max_version_row
                        found_optimal_position = True
                        
                        # Get version info for logging
                        col_c_value = str(worksheet.cell(row=insertion_row, column=3).value or "")
                        v_match = re.search(r'V(\d+)', col_c_value)
                        current_version = int(v_match.group(1)) if v_match else 0
                        
                        self.app.log_message(f"✅ Found optimal position in sheet '{sheet_name}' at row {insertion_row} (suffix: {max_suffix}, version: {current_version} -> {new_version})")
                        break
                
                if not found_optimal_position:
                    # No optimal position found - find the correct sheet and insert at the end
                    target_sheet = None
                    last_data_row = 8
                    
                    # Search for the correct sheet based on prefix
                    for sheet_name in workbook.sheetnames:
                        worksheet = workbook[sheet_name]
                        
                        # Check if this sheet contains entries with the same document ID
                        for row_num in range(9, worksheet.max_row + 1):
                            col_b_value = str(worksheet.cell(row=row_num, column=2).value or "")
                            
                            if self.has_same_document_id(attachment_filename, col_b_value):
                                # Found the correct sheet
                                target_sheet = sheet_name
                                # Find the last row with data in this sheet
                                for check_row in range(9, worksheet.max_row + 1):
                                    if worksheet.cell(row=check_row, column=1).value is not None:
                                        last_data_row = check_row
                                insertion_row = last_data_row
                                self.app.log_message(f"ℹ️ Found correct sheet '{sheet_name}' - inserting at end of data (row {insertion_row})")
                                break
                        
                        if target_sheet:
                            break
                    
                    # If no sheet found with matching prefix, use the first sheet
                    if not target_sheet:
                        target_sheet = workbook.sheetnames[0]
                        worksheet = workbook[target_sheet]
                        # Find the last row with data
                        for row_num in range(9, worksheet.max_row + 1):
                            if worksheet.cell(row=row_num, column=1).value is not None:
                                last_data_row = row_num
                        insertion_row = last_data_row
                        self.app.log_message(f"ℹ️ No matching sheet found - using first sheet '{target_sheet}' - inserting at end of data (row {insertion_row})")
                
                # Get the target worksheet
                worksheet = workbook[target_sheet]
                
                # Insert new row at the determined position
                worksheet.insert_rows(insertion_row + 1)
                
                # Copy formulas and formatting from the row above
                formula_count = 0
                for col_num in range(1, worksheet.max_column + 1):
                    above_cell = worksheet.cell(row=insertion_row, column=col_num)
                    new_cell = worksheet.cell(row=insertion_row + 1, column=col_num)
                    
                    # Copy cell formatting (style, borders, etc.)
                    if above_cell.has_style:
                        new_cell._style = above_cell._style
                    
                    # Check if the above cell contains a formula
                    if above_cell.value and isinstance(above_cell.value, str) and above_cell.value.startswith('='):
                        try:
                            # Manual formula adjustment
                            old_formula = above_cell.value
                            new_formula = self._adjust_formula_for_new_row_intelligent(old_formula, insertion_row, insertion_row + 1, col_num)
                            new_cell.value = new_formula
                            formula_count += 1
                        except Exception as e:
                            pass  # Silently handle formula errors
                    else:
                        # Copy value if it's not a formula
                        new_cell.value = above_cell.value
                    
                    # Copy hyperlink if present
                    if above_cell.hyperlink:
                        try:
                            old_hyperlink = above_cell.hyperlink
                            new_target = self._adjust_hyperlink_target(old_hyperlink.target, insertion_row, insertion_row + 1)
                            new_cell.hyperlink = new_target
                        except Exception as e:
                            pass  # Silently handle hyperlink errors
                
                # Fill in the new row with filename parts
                self._fill_new_row_with_filename_parts(worksheet, insertion_row + 1, attachment_filename, "A", has_multiple_formats, all_files_in_group)
                
                # Set the correct formula for column I in the new row
                self._set_column_i_formula(worksheet, insertion_row + 1)
                
                # Handle hyperlinks for the new row
                self._update_hyperlink_logic(worksheet, insertion_row, attachment_filename, is_v1_file, has_multiple_formats, all_files_in_group)
                
                # Dialog is now handled in _update_hyperlink_logic with priority file check
                
                match_found = True  # Mark as found so it gets saved
            
            # If matches were found, process the last match
            elif all_matches:
                # Get the last match
                last_match = all_matches[-1]
                matched_sheet = last_match['sheet_name']
                last_matched_row = last_match['row_num']
                old_status = last_match['old_status']
                
                # Get the worksheet for the last match
                worksheet = workbook[matched_sheet]
                
                # Enhanced logic for matching and comparison - determine insertion position
                insertion_row = last_matched_row
                
                # Look for "V" character in the filename and check the character beside it
                if is_v1_file:
                    # Compare document IDs in Column B of existing rows
                    for compare_row in range(9, last_matched_row + 1):  # Start from row 9
                        col_b_value = str(worksheet.cell(row=compare_row, column=2).value or "")
                        
                        if self.has_same_document_id(attachment_filename, col_b_value):
                            # Compare last 3 characters as numbers
                            try:
                                current_last3 = int(col_b_value[-3:]) if len(col_b_value) >= 3 else 0
                                new_last3 = int(attachment_filename[7:10]) if len(attachment_filename) >= 10 else 0
                                
                                if new_last3 > current_last3:
                                    # Found a match with larger number - use this position
                                    insertion_row = compare_row
                                    break
                            except ValueError:
                                continue
                
                # Set status in column A to "E" (Executed/Edited) for the matched row
                worksheet.cell(row=insertion_row, column=1, value="E")
                
                # Insert a new row below the determined insertion row
                worksheet.insert_rows(insertion_row + 1)
                
                # Copy formulas and formatting from the row above and adjust them
                formula_count = 0
                for col_num in range(1, worksheet.max_column + 1):
                    above_cell = worksheet.cell(row=insertion_row, column=col_num)
                    new_cell = worksheet.cell(row=insertion_row + 1, column=col_num)
                    
                    # Copy cell formatting (style, borders, etc.)
                    if above_cell.has_style:
                        new_cell._style = above_cell._style
                    
                    # Check if the above cell contains a formula
                    if above_cell.value and isinstance(above_cell.value, str) and above_cell.value.startswith('='):
                        try:
                            # Manual formula adjustment - Excel's automatic adjustment is not working correctly
                            old_formula = above_cell.value
                            new_formula = self._adjust_formula_for_new_row_intelligent(old_formula, insertion_row, insertion_row + 1, col_num)
                            new_cell.value = new_formula
                            formula_count += 1
                        except Exception as e:
                            pass  # Silently handle formula errors
                    else:
                        # Copy value if it's not a formula
                        new_cell.value = above_cell.value
                    
                    # Copy hyperlink if present
                    if above_cell.hyperlink:
                        try:
                            old_hyperlink = above_cell.hyperlink
                            # Create new hyperlink with adjusted target
                            new_target = self._adjust_hyperlink_target(old_hyperlink.target, insertion_row, insertion_row + 1)
                            new_cell.hyperlink = new_target
                        except Exception as e:
                            pass  # Silently handle hyperlink errors
                
                # Now adjust ALL formulas and hyperlinks in rows below the inserted row
                adjusted_formulas_count = 0
                adjusted_hyperlinks_count = 0
                skipped_empty_rows = 0
                
                for adjust_row in range(insertion_row + 2, worksheet.max_row + 1):
                    # Check if column A has a value (only for rows 9 and above)
                    if adjust_row >= 9:
                        column_a_value = worksheet.cell(row=adjust_row, column=1).value
                        if column_a_value is None or str(column_a_value).strip() == "":
                            # Skip this row if column A is empty
                            skipped_empty_rows += 1
                            continue
                    
                    # Process this row since column A has data
                    for col_num in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=adjust_row, column=col_num)
                        
                        # Handle formulas
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            try:
                                old_formula = cell.value
                                new_formula = self._adjust_formula_for_new_row_intelligent(old_formula, insertion_row, adjust_row, col_num)
                                cell.value = new_formula
                                adjusted_formulas_count += 1
                            except Exception as e:
                                pass  # Silently handle formula errors
                        
                        # Handle hyperlinks
                        if cell.hyperlink:
                            try:
                                old_hyperlink = cell.hyperlink
                                # Create new hyperlink with adjusted target
                                new_target = self._adjust_hyperlink_target(old_hyperlink.target, insertion_row, adjust_row)
                                cell.hyperlink = new_target
                                adjusted_hyperlinks_count += 1
                            except Exception as e:
                                pass  # Silently handle hyperlink errors
                
                # Validate and repair formulas after insertion
                self._validate_and_repair_formulas(worksheet, insertion_row)
                
                # Handle hyperlinks for the found row and new row
                self._update_hyperlink_logic(worksheet, insertion_row, attachment_filename, is_v1_file, has_multiple_formats, all_files_in_group)
                
                # Dialog is now handled in _update_hyperlink_logic with priority file check
                
                # Fill in the new row with split filename parts
                self._fill_new_row_with_filename_parts(worksheet, insertion_row + 1, attachment_filename, old_status, has_multiple_formats, all_files_in_group)
                
                # Set the correct formula for column I in the new row
                self._set_column_i_formula(worksheet, insertion_row + 1)
            

            
            # Save changes if a match was found
            if match_found:
                # Force recalculation of all formulas before saving
                try:
                    for sheet_name in workbook.sheetnames:
                        worksheet = workbook[sheet_name]
                        # Force Excel to recalculate formulas
                        worksheet.calculate_dimension()
                except Exception as e:
                    pass  # Silently handle recalculation errors
                
                # Save with data_only=False to preserve formulas
                try:
                    workbook.save(excel_file_path)
                    self.app.log_message(f"✅ Successfully updated Excel tracking file: {Path(excel_file_path).name}")
                    
                    # Verify the file is still valid by trying to load it
                    try:
                        with warnings.catch_warnings():
                            warnings.simplefilter("ignore")
                            test_workbook = load_workbook(excel_file_path, data_only=False)
                            test_workbook.close()
                    except Exception as e:
                        pass  # Silently handle verification errors
                    
                    return True
                except Exception as e:
                    self.app.log_message(f"❌ Error saving workbook: {str(e)}")
                    return False
            else:
                self.app.log_message(f"⚠️ No matching row found for prefix: '{doc_prefix}' and not a V1.0 file")
                self.app.log_message(f"🔍 is_v1_file = {is_v1_file}, filename = {attachment_filename}")
                return False
                
        except Exception as e:
            self.app.log_message(f"❌ Error updating Excel tracking file: {str(e)}")
            return False
        finally:
            if 'workbook' in locals():
                workbook.close()

    def _split_filename_with_regex(self, attachment_filename):
        """
        Split filename into document ID, version-language, and title using regex.
        
        Args:
            attachment_filename (str): The filename to split
            
        Returns:
            tuple: (col_b_value, col_c_value, col_d_value)
        """
        try:
            # Remove file extension if present
            filename_without_ext = attachment_filename
            if '.' in filename_without_ext:
                filename_without_ext = filename_without_ext.rsplit('.', 1)[0]

            # Use regex to extract parts
            # Pattern: Document ID - Version-Language _ Title
            # Example: ABC-DEF-123-V1.0-DE_Some_Title
            pattern = r'^([A-Z]+-[A-Z]+-\d{3})-(V\d+\.\d+-[A-Z]+)_(.+)$'
            match = re.match(pattern, filename_without_ext)

            if match:
                col_b_value = match.group(1)  # Document ID (e.g., "ABC-DEF-123")
                col_c_value = match.group(2)  # Version-Language (e.g., "V1.0-DE")
                col_d_value = match.group(3)  # Title (e.g., "Some_Title")
                self.app.log_message(f"✅ Regex match successful: ID='{col_b_value}', Ver='{col_c_value}', Title='{col_d_value}'")
            else:
                # Fallback if pattern doesn't match
                self.app.log_message(f"⚠️ Regex pattern didn't match filename: '{filename_without_ext}'. Using fallback.")
                col_b_value = filename_without_ext
                col_c_value = ""
                col_d_value = ""

            return col_b_value, col_c_value, col_d_value
            
        except Exception as e:
            self.app.log_message(f"❌ Error in regex filename splitting: {str(e)}")
            # Return filename as-is in case of error
            return attachment_filename, "", ""

    def extract_document_id(self, filename):
        """
        Extract Document ID from filename using regex for exact matching.
        
        Args:
            filename (str): The filename to extract document ID from
            
        Returns:
            str or None: The document ID (e.g., "ABC-DEF-123") or None if not found
        """
        try:
            # Remove extension (if any)
            filename_without_ext = filename.split('.')[0] if '.' in filename else filename
            
            # Extract Document ID (before " - " or at the start matching pattern)
            match = re.match(r"^([A-Z]+-[A-Z]+-\d{3})", filename_without_ext)
            return match.group(1) if match else None
            
        except Exception as e:
            self.app.log_message(f"❌ Error extracting document ID from '{filename}': {str(e)}")
            return None

    def _parse_structured_filename(self, filename):
        """
        Parse structured filename into B, C, D columns using structured parsing.
        Alternative to _split_filename_with_regex with more flexible parsing.
        
        Args:
            filename (str): The filename to parse
            
        Returns:
            tuple: (col_b, col_c, col_d) representing Document ID, Version-Language, Title
        """
        try:
            # Remove extension
            filename_no_ext = filename.split('.')[0] if '.' in filename else filename
            
            # Split into [Document ID] - [Version-Language] _ [Title]
            parts = re.split(r"\s*-\s+|\s*_", filename_no_ext)
            
            # Assign parts to columns
            col_b = parts[0] if len(parts) > 0 else ""  # Document ID
            col_c = parts[1] if len(parts) > 1 else ""  # Version-Language
            col_d = parts[2] if len(parts) > 2 else ""  # Title
            
            self.app.log_message(f"✅ Structured parsing: ID='{col_b}', Ver='{col_c}', Title='{col_d}'")
            return col_b, col_c, col_d
            
        except Exception as e:
            self.app.log_message(f"❌ Error in structured filename parsing: {str(e)}")
            # Return filename as-is in case of error
            return filename, "", ""

    def has_same_document_id(self, filename, cell_value):
        """
        Check if filename and cell value have the same document ID.
        
        Args:
            filename (str): The attachment filename
            cell_value (str): The value from column B in Excel
            
        Returns:
            bool: True if they have the same document ID, False otherwise
        """
        try:
            doc_id_from_filename = self.extract_document_id(filename)
            doc_id_from_cell = self.extract_document_id(str(cell_value).strip()) if cell_value else None
            
            return (doc_id_from_filename is not None and 
                    doc_id_from_cell is not None and 
                    doc_id_from_filename == doc_id_from_cell)
        except Exception:
            return False

    def _fill_new_row_with_filename_parts(self, worksheet, row_num, attachment_filename, old_status, has_multiple_formats=False, all_files_in_group=None):
        """
        Updated: Split filename into document ID, version-language, and title.
        Writes to columns B, C, D.
        
        Args:
            worksheet: The worksheet to modify
            row_num (int): The row number to fill
            attachment_filename (str): The filename to split
            old_status: The status to put in column A
            has_multiple_formats (bool): Whether this file has multiple format versions
            all_files_in_group (list): List of all files in the same group (for multiple formats)
        """
        try:
            # Column A: Old status
            worksheet.cell(row=row_num, column=1, value=old_status)

            # Use the new regex-based splitting
            col_b_value, col_c_value, col_d_value = self._split_filename_with_regex(attachment_filename)

            # Handle multiple formats in column D (title)
            # Note: Multiple formats indicator removed per user request
            # Only the priority file gets inserted into Excel without format indicators

            # Set values
            worksheet.cell(row=row_num, column=2, value=col_b_value)
            worksheet.cell(row=row_num, column=3, value=col_c_value)
            worksheet.cell(row=row_num, column=4, value=col_d_value)
                
        except Exception as e:
            self.app.log_message(f"❌ Error filling new row with filename parts: {str(e)}")

    def _update_hyperlink_logic(self, worksheet, row_num, attachment_filename, is_v1_file=False, has_multiple_formats=False, all_files_in_group=None):
        """
        Update hyperlink logic for the found row and new row.
        
        Args:
            worksheet: The worksheet to modify
            row_num (int): The row number where the match was found
            attachment_filename (str): The filename for creating hyperlinks
            is_v1_file (bool): Whether this is a V1.0 file (no found row to update)
            has_multiple_formats (bool): Whether this file has multiple format versions
            all_files_in_group (list): List of all files in the same group (for multiple formats)
        """
        try:
            # Store current attachment filename for dialogs
            self.current_attachment_filename = attachment_filename
            
            # Get archive and target paths from the app
            archive_dir = self.app.archive_entry.get().strip()
            target_dir = self.app.target_entry.get().strip()
            
            if not archive_dir or not target_dir:
                return
            
            # Determine the priority filename for hyperlinks (PDF > Word > Excel > others)
            hyperlink_filename = attachment_filename  # Default to current file
            if has_multiple_formats and all_files_in_group:
                # Find the priority file for hyperlinks
                pdf_files = [f for f in all_files_in_group if f.suffix.lower() == '.pdf']
                docx_files = [f for f in all_files_in_group if f.suffix.lower() == '.docx']
                xlsx_files = [f for f in all_files_in_group if f.suffix.lower() == '.xlsx']
                
                if pdf_files:
                    hyperlink_filename = pdf_files[0].name
                    self.app.log_message(f"🔗 Using PDF for hyperlinks: {hyperlink_filename}")
                elif docx_files:
                    hyperlink_filename = docx_files[0].name
                    self.app.log_message(f"🔗 Using Word for hyperlinks: {hyperlink_filename}")
                elif xlsx_files:
                    hyperlink_filename = xlsx_files[0].name
                    self.app.log_message(f"🔗 Using Excel for hyperlinks: {hyperlink_filename}")
                else:
                    hyperlink_filename = all_files_in_group[0].name
                    self.app.log_message(f"🔗 Using first file for hyperlinks: {hyperlink_filename}")
            
            if not is_v1_file:
                # 1. Replace the old hyperlink in the found cell (column J) with archive dir + priority file
                found_row_cell_j = worksheet.cell(row=row_num, column=10)  # Column J is index 10
                archive_path = Path(archive_dir) / hyperlink_filename
                
                if archive_path.exists():
                    archive_hyperlink = f"file:///{archive_path.absolute().as_posix()}"
                    # Show the full archive path as the hyperlink text
                    found_row_cell_j.value = str(archive_path)
                    found_row_cell_j.hyperlink = archive_hyperlink
                else:
                    self.app.log_message(f"⚠️ Archive document not found at: {archive_path}")
                    found_row_cell_j.value = f"{archive_path} (Not Found)"
            
            # 2. Add hyperlink in the new row (column J) with target dir + priority file
            new_row_cell_j = worksheet.cell(row=row_num + 1, column=10)  # Column J is index 10
            target_path = Path(target_dir) / hyperlink_filename
            
            if target_path.exists():
                target_hyperlink = f"file:///{target_path.absolute().as_posix()}"
                
                # Show the full target path as the hyperlink text (no format indicators)
                new_row_cell_j.value = str(target_path)
                
                new_row_cell_j.hyperlink = target_hyperlink
            else:
                self.app.log_message(f"⚠️ Target document not found at: {target_path}")
                new_row_cell_j.value = f"{target_path} (Not Found)"
            
            # 3. Show dialog for user input in columns E, F, G
            # Only show dialog for priority files (PDF) in duplicate groups
            should_show_dialog = True
            
            # Check if this is a duplicate group and if this file is the priority file
            if has_multiple_formats and all_files_in_group:
                # Find the priority file (PDF > Word > Excel > others)
                pdf_files = [f for f in all_files_in_group if f.suffix.lower() == '.pdf']
                if pdf_files:
                    priority_file = pdf_files[0]
                    # Only show dialog if this is the priority file
                    should_show_dialog = (priority_file.name == attachment_filename)
                    if not should_show_dialog:
                        self.app.log_message(f"⏭️ Skipping dialog in _update_hyperlink_logic for non-priority file: {attachment_filename}")
            
            if should_show_dialog:
                if is_v1_file:
                    self._show_v1_cell_input_dialog(worksheet, row_num, has_multiple_formats, all_files_in_group)
                else:
                    self._show_cell_input_dialog(worksheet, row_num, has_multiple_formats, all_files_in_group)
            
        except Exception as e:
            self.app.log_message(f"❌ Error updating hyperlink logic: {str(e)}")
    
    def _show_v1_cell_input_dialog(self, worksheet, row_num, has_multiple_formats=False, all_files_in_group=None):
        """
        Show dialog for user input in columns E, F, G for V1.0 files.
        V1.0 files only need new row data (no found row to update).
        
        Args:
            worksheet: The worksheet to modify
            row_num (int): The row number where the new row was inserted
            has_multiple_formats (bool): Whether this file has multiple format versions
            all_files_in_group (list): List of all files in the same group (for multiple formats)
        """
        try:
            # For V1.0 files, we only need to set up the new row data
            new_row_data = {
                'E': '',              # Empty date for new entry
                'F': 'aktuell gültig', # Fixed value for new row
                'G': '-'              # Fixed value for new row
            }
            
            # Prepare document information
            document_info = {
                'filename': getattr(self, 'current_attachment_filename', 'Unknown'),
                'doc_prefix': getattr(self, 'current_doc_prefix', 'Unknown'),
                'sheet_name': worksheet.title,
                'row_num': row_num,
                'is_v1_file': True,
                'has_multiple_formats': has_multiple_formats,
                'all_files_in_group': all_files_in_group
            }
            
            # Show the dialog with only new row data (no found row)
            try:
                self.app.log_message(f"🔍 Creating V1.0 Excel cell input dialog for: {document_info.get('filename', 'Unknown')}")
                dialog = ExcelCellInputDialog(self.app.root, None, new_row_data, document_info)
                # Ensure dialog is properly shown
                self.app.log_message(f"🔍 Showing V1.0 Excel cell input dialog...")
                dialog.show_dialog()
                self.app.root.wait_window(dialog.dialog)
                self.app.log_message(f"🔍 V1.0 Excel cell input dialog closed")
            except Exception as e:
                self.app.log_message(f"❌ Error creating V1.0 Excel dialog: {str(e)}")
                # Fallback: use default values
                dialog = None
                dialog.result = {'new_row': {'E': '', 'F': 'aktuell gültig', 'G': '-'}}
            
            # Apply the user input if dialog was confirmed
            if dialog and dialog.result:
                # Apply new row values only (no found row to update)
                for col, value in dialog.result['new_row'].items():
                    col_index = {'E': 5, 'F': 6, 'G': 7}[col]
                    # Parse date if it's in DD.MM.YYYY format
                    parsed_value = self._parse_date_value(value)
                    worksheet.cell(row=row_num + 1, column=col_index, value=parsed_value)
                
                self.app.log_message(f"✅ Successfully applied user input for V1.0 file columns E, F, G")
            else:
                self.app.log_message(f"⚠️ User cancelled V1.0 cell input dialog")
            
        except Exception as e:
            self.app.log_message(f"❌ Error showing V1.0 cell input dialog: {str(e)}")
    
    def _show_cell_input_dialog(self, worksheet, row_num, has_multiple_formats=False, all_files_in_group=None):
        """
        Show dialog for user input in columns E, F, G for both found row and new row.
        The new row will have the exact same format as the found row.
        
        Args:
            worksheet: The worksheet to modify
            row_num (int): The row number where the match was found
            has_multiple_formats (bool): Whether this file has multiple format versions
            all_files_in_group (list): List of all files in the same group (for multiple formats)
        """
        try:
            # Get current values from found row
            found_row_data = {
                'E': worksheet.cell(row=row_num, column=5).value or '',  # Column E is index 5
                'F': worksheet.cell(row=row_num, column=6).value or '',  # Column F is index 6
                'G': worksheet.cell(row=row_num, column=7).value or ''   # Column G is index 7
            }
            
            # Copy the exact same format from found row to new row, but fix F and G
            new_row_data = {
                'E': found_row_data['E'],  # Copy exact same value for date format
                'F': 'aktuell gültig',     # Fixed value for new row
                'G': '-'                   # Fixed value for new row
            }
            
            # Prepare document information
            document_info = {
                'filename': getattr(self, 'current_attachment_filename', 'Unknown'),
                'doc_prefix': getattr(self, 'current_doc_prefix', 'Unknown'),
                'sheet_name': worksheet.title,
                'row_num': row_num,
                'is_v1_file': False,
                'has_multiple_formats': has_multiple_formats,
                'all_files_in_group': all_files_in_group
            }
            
            # Show the dialog
            try:
                self.app.log_message(f"🔍 Creating regular Excel cell input dialog for: {document_info.get('filename', 'Unknown')}")
                dialog = ExcelCellInputDialog(self.app.root, found_row_data, new_row_data, document_info)
                # Ensure dialog is properly shown
                self.app.log_message(f"🔍 Showing regular Excel cell input dialog...")
                dialog.show_dialog()
                self.app.root.wait_window(dialog.dialog)
                self.app.log_message(f"🔍 Regular Excel cell input dialog closed")
            except Exception as e:
                self.app.log_message(f"❌ Error creating Excel dialog: {str(e)}")
                # Fallback: use default values
                dialog = None
                dialog.result = {'new_row': {'E': '', 'F': 'aktuell gültig', 'G': '-'}}
            
            # Apply the user input if dialog was confirmed
            if dialog and dialog.result:
                # Apply found row values (only if they exist)
                if 'found_row' in dialog.result:
                    for col, value in dialog.result['found_row'].items():
                        col_index = {'E': 5, 'F': 6, 'G': 7}[col]
                        # Parse date if it's in DD.MM.YYYY format
                        parsed_value = self._parse_date_value(value)
                        worksheet.cell(row=row_num, column=col_index, value=parsed_value)
                
                # Apply new row values (with exact same format as found row)
                for col, value in dialog.result['new_row'].items():
                    col_index = {'E': 5, 'F': 6, 'G': 7}[col]
                    # Parse date if it's in DD.MM.YYYY format
                    parsed_value = self._parse_date_value(value)
                    worksheet.cell(row=row_num + 1, column=col_index, value=parsed_value)
                
                self.app.log_message(f"✅ Successfully applied user input for columns E, F, G")
            else:
                self.app.log_message(f"⚠️ User cancelled cell input dialog")
            
        except Exception as e:
            self.app.log_message(f"❌ Error showing cell input dialog: {str(e)}")

    def _parse_date_value(self, value):
        """
        Parse a date value from DD.MM.YYYY format to a proper Excel date object.
        
        Args:
            value (str): The date string in DD.MM.YYYY format
            
        Returns:
            datetime.date or str: Parsed date object or original value if not a date
        """
        try:
            if not value or not isinstance(value, str):
                return value
            
            # Check if the value matches DD.MM.YYYY format
            import re
            date_pattern = r'^\d{2}\.\d{2}\.\d{4}$'
            if re.match(date_pattern, value.strip()):
                # Parse the date
                from datetime import datetime
                parsed_date = datetime.strptime(value.strip(), "%d.%m.%Y")
                return parsed_date.date()  # Return date object without time component
            else:
                # Not a date, return original value
                return value
                
        except Exception as e:
            self.app.log_message(f"⚠️ Error parsing date value '{value}': {str(e)}")
            return value

    def _set_column_i_formula(self, worksheet, row_num):
        """
        Set the correct formula for column I in the new row.
        This should be a CONCATENATE formula that combines the values from columns B, C, and D.
        
        Args:
            worksheet: The worksheet to modify
            row_num (int): The row number to set the formula in
        """
        try:
            # Create the CONCATENATE formula for column I
            # This combines the values from columns B, C, and D with appropriate separators
            column_i_formula = f'=CONCATENATE(B{row_num},"-",C{row_num},"_",D{row_num})'
            
            # Set the formula in column I
            worksheet.cell(row=row_num, column=9, value=column_i_formula)  # Column I is index 9
            
        except Exception as e:
            self.app.log_message(f"❌ Error setting column I formula: {str(e)}")

    def _adjust_formula_for_new_row_intelligent(self, formula, old_row, new_row, current_col):
        """
        Intelligently adjust Excel formula for a new row by updating row references.
        This method handles table references properly and avoids corrupting them.
        
        Args:
            formula (str): The original formula
            old_row (int): The original row number (insertion point)
            new_row (int): The new row number (insertion point + 1)
            current_col (int): The current column number (for context)
            
        Returns:
            str: The adjusted formula
        """
        try:
            import re
            
            # Check if this is a table reference formula (like Tabelle142514[@Kürzel])
            if '[@' in formula:
                return formula
            
            # Pattern to match cell references (e.g., A1, B2, C10, etc.)
            # But exclude table references like Tabelle142514[@Kürzel]
            cell_pattern = r'([A-Z]+)(\d+)'
            
            def replace_cell_ref(match):
                column = match.group(1)
                row = int(match.group(2))
                
                # Check if this is part of a table reference (should not be adjusted)
                full_match = match.group(0)
                start_pos = match.start()
                end_pos = match.end()
                
                # Look for table reference patterns around this match
                formula_before = formula[:start_pos]
                formula_after = formula[end_pos:]
                
                # If we find table reference patterns, don't adjust this cell reference
                if re.search(r'\[@[^\]]*\]', formula_before + formula_after):
                    return full_match
                
                # ALL rows at or below the insertion point should be incremented by 1
                # This is because inserting a row shifts all subsequent rows down
                if row >= old_row:
                    new_row_num = row + 1
                    new_ref = f"{column}{new_row_num}"
                    return new_ref
                else:
                    # Keep row references above the insertion point unchanged
                    return full_match
            
            # Apply the replacement
            adjusted_formula = re.sub(cell_pattern, replace_cell_ref, formula)
            
            return adjusted_formula
            
        except Exception as e:
            # Return original formula if adjustment fails
            return formula

    def _adjust_hyperlink_target(self, target, old_row, new_row):
        """
        Adjust hyperlink target for a new row by updating row references.
        
        Args:
            target (str): The original hyperlink target (e.g., "Sheet1!A26", "#Sheet1!A26", etc.)
            old_row (int): The original row number (insertion point)
            new_row (int): The new row number
            
        Returns:
            str: The adjusted hyperlink target
        """
        try:
            import re
            
            # Pattern to match cell references in hyperlink targets
            # This handles various formats like "Sheet1!A26", "#Sheet1!A26", "A26", etc.
            cell_pattern = r'([A-Z]+)(\d+)'
            
            def replace_cell_ref(match):
                column = match.group(1)
                row = int(match.group(2))
                
                # ALL rows at or below the insertion point should be incremented by 1
                if row >= old_row:
                    new_row_num = row + 1
                    new_ref = f"{column}{new_row_num}"
                    return new_ref
                else:
                    # Keep row references above the insertion point unchanged
                    return match.group(0)
            
            # Apply the replacement
            adjusted_target = re.sub(cell_pattern, replace_cell_ref, target)
            
            return adjusted_target
            
        except Exception as e:
            # Return original target if adjustment fails
            return target

    def _adjust_formula_for_new_row(self, formula, old_row, new_row):
        """
        Adjust Excel formula for a new row by updating row references.
        
        Args:
            formula (str): The original formula
            old_row (int): The original row number
            new_row (int): The new row number
            
        Returns:
            str: The adjusted formula
        """
        try:
            import re
            
            # Simple approach: replace row numbers in cell references
            # This handles basic cases like A1, B2, etc.
            cell_pattern = r'([A-Z]+)(\d+)'
            
            def replace_cell_ref(match):
                column = match.group(1)
                row = int(match.group(2))
                
                # If this cell reference is in the same row as the original, adjust it
                if row == old_row:
                    new_ref = f"{column}{new_row}"
                    return new_ref
                else:
                    # Keep other cell references unchanged
                    return match.group(0)
            
            # Apply the replacement
            adjusted_formula = re.sub(cell_pattern, replace_cell_ref, formula)
            
            return adjusted_formula
            
        except Exception as e:
            # Return original formula if adjustment fails
            return formula

    def _validate_and_repair_formulas(self, worksheet, insertion_row):
        """
        Validate and repair formulas after row insertion to ensure they display correctly.
        
        Args:
            worksheet: The worksheet to validate
            insertion_row (int): The row where insertion occurred
        """
        try:
            repaired_count = 0
            
            # Check all cells in the worksheet for formulas
            for row in range(1, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        try:
                            # Try to validate the formula by checking if it's syntactically correct
                            formula = cell.value
                            
                            # Check for common formula issues after row insertion
                            if '#REF!' in formula or '#N/A' in formula:
                                # Try to repair the formula
                                repaired_formula = self._repair_formula(formula, insertion_row)
                                if repaired_formula != formula:
                                    cell.value = repaired_formula
                                    repaired_count += 1
                            
                        except Exception as e:
                            pass
            
            if repaired_count > 0:
                self.app.log_message(f"🔧 Formula validation complete. Repaired {repaired_count} formulas.")
            
        except Exception as e:
            pass

    def _repair_formula(self, formula, insertion_row):
        """
        Attempt to repair a formula that may have been corrupted during row insertion.
        
        Args:
            formula (str): The formula to repair
            insertion_row (int): The row where insertion occurred
            
        Returns:
            str: The repaired formula
        """
        try:
            import re
            
            # Pattern to match cell references
            cell_pattern = r'([A-Z]+)(\d+)'
            
            def repair_cell_ref(match):
                column = match.group(1)
                row = int(match.group(2))
                
                # If the row reference is at or below the insertion point, it should be incremented
                if row >= insertion_row:
                    new_row = row + 1
                    return f"{column}{new_row}"
                else:
                    return match.group(0)
            
            # Apply the repair
            repaired_formula = re.sub(cell_pattern, repair_cell_ref, formula)
            
            return repaired_formula
            
        except Exception as e:
            return formula

    def add_watermark_to_excel(self, file_path, watermark_text="UNGÜLTIG", font_size=36, font_color="FF0000", transparency=0.7):
        """
        Add diagonal red text field watermark to Excel file using Excel automation
        This creates actual floating text boxes exactly like manual insertion
        
        Args:
            file_path (str): Path to the Excel document
            watermark_text (str): Watermark text (default: "UNGÜLTIG")
            font_size (int): Font size in points (default: 36)
            font_color (str): Color in hex format (default: "FF0000" for red)
            transparency (float): Transparency level 0.0-1.0 (default: 0.7)
        """
        try:
            # Priority 1: Use win32com for Excel automation (creates actual floating text boxes)
            if WIN32COM_AVAILABLE:
                try:
                    return self._add_diagonal_text_field_with_win32com(file_path, watermark_text, font_size, font_color, transparency)
                except Exception as win32com_error:
                    self.app.log_message(f"⚠️ win32com failed, trying Spire.XLS: {str(win32com_error)}")
            
            # Priority 2: Use Spire.XLS for advanced watermarking with diagonal text field
            if SPIRE_XLS_AVAILABLE:
                try:
                    return self._add_diagonal_text_field_with_spire(file_path, watermark_text, font_size, font_color, transparency)
                except Exception as spire_error:
                    self.app.log_message(f"⚠️ Spire.XLS failed, falling back to openpyxl: {str(spire_error)}")
            
            # Priority 3: Fall back to openpyxl watermarking (cell-based, not floating text boxes)
            self.app.log_message(f"⚠️ Using fallback method (cell-based watermarking)")
            return self._add_watermark_with_openpyxl(file_path, watermark_text, font_size, font_color, transparency)
                
        except Exception as e:
            self.app.log_message(f"❌ Error adding watermark to Excel: {str(e)}")
            return False

    def _add_image_watermark_with_openpyxl(self, file_path, watermark_image_path):
        """
        Add image watermark to Excel file using openpyxl
        """
        try:
            # Load the workbook
            workbook = load_workbook(file_path)
            
            # Add image watermark to each worksheet
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                self._add_image_watermark_to_worksheet(worksheet, watermark_image_path)
            
            # Save the watermarked workbook
            workbook.save(file_path)
            workbook.close()
            
            self.app.log_message(f"✅ Image watermark added to {Path(file_path).name}")
            return True
            
        except Exception as e:
            self.app.log_message(f"❌ Error adding image watermark with openpyxl: {str(e)}")
            return False

    def _add_image_watermark_to_worksheet(self, worksheet, watermark_image_path):
        """
        Add image watermark to a specific worksheet
        """
        try:
            # Get worksheet dimensions
            max_row = worksheet.max_row or 1
            max_col = worksheet.max_column or 1
            
            # Calculate center position
            center_row = max_row // 2
            center_col = max_col // 2
            
            # Create image object
            img = Image(str(watermark_image_path))
            
            # Set image size (make it large enough to be visible)
            img.width = 300
            img.height = 100
            
            # Position the image in the center of the worksheet
            worksheet.add_image(img, f'{get_column_letter(center_col)}{center_row}')
            
        except Exception as e:
            self.app.log_message(f"⚠️ Error adding image watermark to worksheet {worksheet.title}: {str(e)}")

    def _add_watermark_with_openpyxl(self, file_path, watermark_text, font_size, font_color, transparency):
        """
        Fallback method using openpyxl for text-based Excel watermarking on ALL sheets
        """
        try:
            # Load the workbook
            workbook = load_workbook(file_path)
            
            # Add watermark to each worksheet
            sheet_count = len(workbook.sheetnames)
            watermarked_sheets = 0
            
            for sheet_name in workbook.sheetnames:
                try:
                    worksheet = workbook[sheet_name]
                    self._add_watermark_to_worksheet(worksheet, watermark_text, font_size, font_color, transparency)
                    watermarked_sheets += 1
                except Exception as sheet_error:
                    self.app.log_message(f"⚠️ Error watermarking sheet '{sheet_name}': {str(sheet_error)}")
                    continue
            
            # Save the watermarked workbook
            workbook.save(file_path)
            workbook.close()
            
            self.app.log_message(f"✅ Text watermark added to {watermarked_sheets}/{sheet_count} sheets in {Path(file_path).name}")
            return True
            
        except Exception as e:
            self.app.log_message(f"❌ Error adding text watermark with openpyxl: {str(e)}")
            return False

    def _add_watermark_to_worksheet(self, worksheet, watermark_text, font_size, font_color, transparency):
        """
        Add text watermark to a specific worksheet using cell styling
        """
        try:
            # Add watermark as text in a cell for visibility
            self._add_watermark_cell(worksheet, watermark_text, font_size, font_color, transparency)
            
        except Exception as e:
            self.app.log_message(f"⚠️ Error adding watermark to worksheet {worksheet.title}: {str(e)}")

    def _add_watermark_cell(self, worksheet, watermark_text, font_size, font_color, transparency):
        """
        Add watermark as text in a cell with special styling
        """
        try:
            # Get worksheet dimensions
            max_row = worksheet.max_row or 1
            max_col = worksheet.max_column or 1
            
            # Calculate center position
            center_row = max_row // 2
            center_col = max_col // 2
            
            # Add watermark text to center cell
            watermark_cell = worksheet.cell(row=center_row, column=center_col)
            watermark_cell.value = watermark_text
            
            # Style the watermark cell
            watermark_font = OpenpyxlFont(
                name='Arial',
                size=font_size,
                color=font_color,
                bold=True
            )
            watermark_cell.font = watermark_font
            
            # Set alignment to center
            watermark_cell.alignment = Alignment(
                horizontal='center',
                vertical='center'
            )
            
            # Merge cells to make watermark more prominent
            try:
                worksheet.merge_cells(
                    start_row=center_row,
                    start_column=center_col,
                    end_row=center_row + 2,
                    end_column=center_col + 2
                )
            except Exception as e:
                self.app.log_message(f"⚠️ Could not merge cells for watermark: {str(e)}")
            
        except Exception as e:
            self.app.log_message(f"⚠️ Error adding watermark cell: {str(e)}")

    def _add_diagonal_text_field_with_spire(self, file_path, watermark_text="UNGÜLTIG", font_size=36, font_color="FF0000", transparency=0.7):
        """
        Add diagonal red text field watermark to Excel file using Spire.XLS on ALL sheets
        
        Args:
            file_path (str): Path to the Excel document
            watermark_text (str): Watermark text (default: "UNGÜLTIG")
            font_size (int): Font size in points (default: 36)
            font_color (str): Color in hex format (default: "FF0000" for red)
            transparency (float): Transparency level 0.0-1.0 (default: 0.7)
        """
        try:
            # Create a Workbook and load the existing Excel file
            workbook = Workbook()
            workbook.LoadFromFile(file_path)
            
            # Process each worksheet in the workbook
            sheet_count = workbook.Worksheets.Count
            watermarked_sheets = 0
            
            for sheet_index in range(sheet_count):
                try:
                    sheet = workbook.Worksheets[sheet_index]
                    
                    # Add a rectangle shape to act as a text box
                    shape = sheet.Shapes.AddShape(AutoShapeType.Rectangle, 100, 100, 300, 100)
                    
                    # Set the text content
                    shape.Text = watermark_text
                    
                    # Format the text using the correct approach
                    format = shape.TextFrame.TextRange.CharacterFormat
                    format.FontName = "Arial"
                    format.Size = font_size
                    format.TextColor = Color.Red
                    format.Bold = True
                    
                    # Rotate the entire shape diagonally (-45 degrees)
                    shape.Rotation = -45
                    
                    # Make background transparent (like a watermark)
                    shape.Fill.FillType = ShapeFillType.None_
                    shape.Line.Visible = False  # Hide the border
                    
                    watermarked_sheets += 1
                    
                except Exception as sheet_error:
                    self.app.log_message(f"⚠️ Error watermarking sheet {sheet_index + 1}: {str(sheet_error)}")
                    continue
            
            # Save the modified Excel file
            workbook.SaveToFile(file_path, ExcelVersion.Version2016)
            
            # Dispose resources
            workbook.Dispose()
            
            self.app.log_message(f"✅ Diagonal red text field watermark added to {watermarked_sheets}/{sheet_count} sheets in {Path(file_path).name}")
            return True
            
        except Exception as e:
            self.app.log_message(f"❌ Error adding diagonal text field watermark with Spire.XLS: {str(e)}")
            return False

    def _add_diagonal_text_field_with_win32com(self, file_path, watermark_text="UNGÜLTIG", font_size=36, font_color="FF0000", transparency=0.7):
        """
        Add diagonal red text field watermark to Excel file using win32com (Excel automation)
        This creates actual floating text boxes exactly like manual insertion on ALL sheets
        
        Args:
            file_path (str): Path to the Excel document
            watermark_text (str): Watermark text (default: "UNGÜLTIG")
            font_size (int): Font size in points (default: 36)
            font_color (str): Color in hex format (default: "FF0000" for red)
            transparency (float): Transparency level 0.0-1.0 (default: 0.7)
        """
        try:
            # Create Excel application object
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False  # Run in background
            excel.DisplayAlerts = False  # Suppress alerts
            
            # Open the workbook
            workbook = excel.Workbooks.Open(os.path.abspath(file_path))
            
            # Process each worksheet in the workbook
            sheet_count = workbook.Worksheets.Count
            watermarked_sheets = 0
            
            for sheet_index in range(1, sheet_count + 1):
                try:
                    worksheet = workbook.Worksheets(sheet_index)
                    
                    # Get worksheet dimensions for positioning
                    used_range = worksheet.UsedRange
                    max_row = used_range.Rows.Count
                    max_col = used_range.Columns.Count
                    
                    # Calculate center position
                    center_row = max_row // 2
                    center_col = max_col // 2
                    
                    # Add a text box shape
                    textbox = worksheet.Shapes.AddTextbox(
                        Orientation=1,  # Horizontal text
                        Left=center_col * 50,  # Position from left
                        Top=center_row * 20,   # Position from top
                        Width=300,             # Width of text box
                        Height=100             # Height of text box
                    )
                    
                    # Set the text content
                    textbox.TextFrame.Characters().Text = watermark_text
                    
                    # Format the text
                    textbox.TextFrame.Characters().Font.Name = "Arial"
                    textbox.TextFrame.Characters().Font.Size = font_size
                    textbox.TextFrame.Characters().Font.Color = 255  # Red color
                    textbox.TextFrame.Characters().Font.Bold = True
                    
                    # Rotate the text box diagonally (-45 degrees)
                    textbox.Rotation = -45
                    
                    # Make background transparent
                    textbox.Fill.Visible = False
                    
                    # Hide the border
                    textbox.Line.Visible = False
                    
                    watermarked_sheets += 1
                    
                except Exception as sheet_error:
                    self.app.log_message(f"⚠️ Error watermarking sheet {sheet_index}: {str(sheet_error)}")
                    continue
            
            # Save the workbook
            workbook.Save()
            
            # Close workbook and quit Excel
            workbook.Close(SaveChanges=True)
            excel.Quit()
            
            self.app.log_message(f"✅ Diagonal red text field watermark added to {watermarked_sheets}/{sheet_count} sheets in {Path(file_path).name} using Excel automation")
            return True
            
        except Exception as e:
            self.app.log_message(f"❌ Error adding diagonal text field watermark with win32com: {str(e)}")
            # Try to clean up Excel if it's still running
            try:
                if 'excel' in locals():
                    excel.Quit()
            except:
                pass
            return False

    def add_watermark_to_archived_excel_files(self, files_to_archive, archive_dir):
        """Add watermarks to Excel files before archiving"""
        processed = []
        for file_path in files_to_archive:
            file_path = Path(file_path)
            if self.is_excel_file(file_path):
                success = self.add_watermark_to_excel(str(file_path))
                status = "✅" if success else "❌"
                self.app.log_message(f"{status} {file_path.name}")
                processed.append(file_path)
        return processed 